# flake8: noqa E501

import io
import os
from datetime import datetime

from flask import (
    Blueprint,
    render_template,
    render_template_string,
    request,
    flash,
    send_file,
)

from flask_login import login_required, current_user
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sqlalchemy as sa

from app.controllers import (
    create_pagination,
    role_required,
    create_ship_requests_by_address,
    save_exel_file,
)

from app import models as m, db
from app import schema as s
from app import forms as f
from app.controllers import validate_bulk_ship_exel
from app.logger import log
from .utils import RELOAD_PAGE_SCRIPT, create_match


bulk_ship_bp = Blueprint("bulk_ship", __name__, url_prefix="/bulk-ship")


@bulk_ship_bp.route("/", methods=["GET"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def get_all():

    q = request.args.get("q", type=str, default="")

    where_stmt = sa.and_(
        m.BulkShip.is_deleted.is_(False), m.BulkShip.user_id == current_user.id
    )

    if current_user.role_obj.role_name == s.UserRole.ADMIN.value:
        where_stmt = sa.and_(m.BulkShip.is_deleted.is_(False))

    if q:
        where_stmt = sa.and_(
            where_stmt,
            sa.or_(
                m.BulkShip.name.ilike(f"%{q}%"),
            ),
        )

    query = (
        sa.select(m.BulkShip).where(where_stmt).order_by(m.BulkShip.created_at.desc())
    )
    count_query = sa.select(sa.func.count()).where(where_stmt).select_from(m.BulkShip)

    pagination = create_pagination(total=db.session.scalar(count_query))
    bulk_ships = db.session.scalars(
        query.offset((pagination.page - 1) * pagination.per_page).limit(
            pagination.per_page
        )
    )

    return render_template(
        "bulk_ship/bulk_ships.html",
        bulk_ships=bulk_ships,
        page=pagination,
        q=q,
    )


@bulk_ship_bp.route("/download-template", methods=["GET"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def download_template():

    wb = Workbook()
    main_sheet = wb.active
    main_sheet.title = "Main"

    # Create the product sheet
    products = db.session.scalars(sa.select(m.Product))
    product_count = db.session.scalar(sa.select(sa.func.count()).select_from(m.Product))
    groups_sheet = wb.create_sheet(title="Groups")
    for col, product in enumerate(products, start=1):
        col_letter = get_column_letter(col)
        groups_sheet[f"{col_letter}1"] = product.SKU
        wh_products = [
            wh_prod
            for wh_prod in product.warehouse_products
            if wh_prod.available_quantity != 0
            and wh_prod.group.name != s.Events.name.value
        ]
        for row, wh_product in enumerate(wh_products, start=2):
            groups_sheet[f"{col_letter}{row}"] = (
                f"{wh_product.group.name} ({wh_product.available_quantity})"
            )

    store_categories = db.session.scalars(
        sa.select(m.StoreCategory).where(m.StoreCategory.name.not_ilike("%Event%"))
    )
    count_store_categories = db.session.scalar(
        sa.select(sa.func.count()).select_from(m.StoreCategory)
    )

    # Create the store categories sheet
    store_categories_sheet = wb.create_sheet(title="Stores")
    for col, store_category in enumerate(store_categories, start=1):
        col_letter = get_column_letter(col)
        store_categories_sheet[f"{col_letter}1"] = store_category.name
        for row, store in enumerate(store_category.stores, start=2):
            store_categories_sheet[f"{col_letter}{row}"] = store.store_name

    # Set up the main form sheet
    main_sheet["A1"] = "SKU"
    main_sheet["B1"] = "Group"
    main_sheet["C1"] = "Quantity"
    main_sheet["D1"] = "Store Category"
    main_sheet["E1"] = "Store name"

    # Add dropdowns for product SKUs
    last_let = get_column_letter(product_count)
    country_dv = DataValidation(
        type="list",
        formula1=f"{groups_sheet.title}!$A$1:${last_let}$1",
        allow_blank=True,
        showDropDown=False,
    )
    main_sheet.add_data_validation(country_dv)
    country_dv.add("A2:A100")

    # Add dropdowns for store category
    store_last_let = get_column_letter(count_store_categories)
    store_category_dv = DataValidation(
        type="list",
        formula1=f"{store_categories_sheet.title}!$A$1:${store_last_let}$1",
        allow_blank=True,
        showDropDown=False,
    )
    main_sheet.add_data_validation(store_category_dv)
    store_category_dv.add("D2:D100")

    # Create data validation for cities using named ranges
    # offset( reference, rows, cols, height, width)
    # OFFSET(Data!A1,1,MATCH(A2,Data!A1:C1,0) -1,10,1)"

    # Create the named ranges
    for row in range(2, 100):
        store_category_dv = DataValidation(
            type="list",
            formula1=f"OFFSET({store_categories_sheet.title}!A1,1,{create_match('D' + str(row), store_categories_sheet.title, count_store_categories)},500,1)",
            showDropDown=False,
            allow_blank=True,
        )
        main_sheet.add_data_validation(store_category_dv)
        store_category_dv.add(f"E{row}")

        groups_dv = DataValidation(
            type="list",
            formula1=f"OFFSET({groups_sheet.title}!A1,1,{create_match('A' + str(row),groups_sheet.title, product_count)},50,1)",
            showDropDown=False,
            allow_blank=False,
            showErrorMessage=True,
        )
        groups_dv.error = "Please select a valid group"
        groups_dv.errorTitle = "Invalid group"
        main_sheet.add_data_validation(groups_dv)
        groups_dv.add(f"B{row}")

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Step 7: Serve the Excel file as a download
    return send_file(
        output,
        as_attachment=True,
        download_name="empty_form_with_dropdowns.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bulk_ship_bp.route("/create", methods=["GET"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def get_create_modal():
    """htmx"""
    form = f.NewBulkShipForm()
    return render_template("bulk_ship/modal_add.html", form=form)


@bulk_ship_bp.route("/create", methods=["POST"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def create():
    """htmx"""
    form = f.NewBulkShipForm()
    if not form.validate_on_submit():
        log(log.ERROR, "Form validation failed", form.errors)

        return render_template(
            "bulk_ship/modal_add.html", form=form, errors=form.errors
        )

    file = request.files.get("exel_file")

    if not file:
        log(log.ERROR, "Form validation failed", form.errors)
        return render_template(
            "bulk_ship/modal_add.html", form=form, errors={"file": ["File is required"]}
        )

    result = s.ValidateBulkShipResult(errors=dict())
    file.seek(0)

    # why we forst save the file and then validate it because after validation file change and save not work properly
    absolute_file_path, upload_file_path = save_exel_file(file, result)
    if result.errors:
        log(log.ERROR, "Exel save failed", result.errors)
        return render_template(
            "bulk_ship/modal_add.html", form=form, errors=result.errors
        )

    wh_products = validate_bulk_ship_exel(file, result)

    if result.errors:
        log(log.ERROR, "Exel validation failed", result.errors)

        os.remove(absolute_file_path)

        for stor_id in result.new_stores_ids:
            store = db.session.get(m.Store, int(stor_id))
            if store:
                db.session.delete(store)

            db.session.commit()
        return render_template(
            "bulk_ship/modal_add.html", form=form, errors=result.errors
        )

    try:
        create_ship_requests_by_address(wh_products)
    except Exception as e:
        log(log.ERROR, "Bulk ship creation failed", str(e))
        return render_template(
            "toast.html", message="Bulk ship creation failed", category="error"
        )

    # TODO mybe we need to notify warehouse manager about new bulk ship

    m.BulkShip(
        user_id=current_user.id,
        status=s.BulkShipStatus.DRAFT.value,
        name=form.name.data,
        absolute_file_path=str(absolute_file_path),
        uploaded_file_path=str(upload_file_path),
    ).save()

    flash("Bulk ship created successfully", category="success")

    return render_template_string(RELOAD_PAGE_SCRIPT)


@bulk_ship_bp.route("/<uuid>/download-template", methods=["GET"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def bulk_ship_template_download(uuid: str):

    bulk_ship = db.session.scalar(sa.select(m.BulkShip).where(m.BulkShip.uuid == uuid))
    if not bulk_ship or bulk_ship.is_deleted:
        log(log.ERROR, "Bulk ship not found [%s]", uuid)
        flash("Bulk ship not found", category="danger")
        return "", 404

    if not os.path.exists(bulk_ship.absolute_file_path):
        log(log.ERROR, "File not found [%s]", bulk_ship.absolute_file_path)
        flash("File not found", category="danger")
        return "", 404

    return send_file(
        bulk_ship.uploaded_file_path,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bulk_ship_bp.route("/<uuid>/view", methods=["GET"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def view(uuid: str):
    """htmx"""
    bulk_ship = db.session.scalar(sa.select(m.BulkShip).where(m.BulkShip.uuid == uuid))
    if not bulk_ship or bulk_ship.is_deleted:
        log(log.ERROR, "Bulk ship not found [%s]", uuid)
        flash("Bulk ship not found", category="danger")
        return render_template_string(RELOAD_PAGE_SCRIPT)

    if not os.path.exists(bulk_ship.absolute_file_path):
        log(log.ERROR, "File not found [%s]", bulk_ship.absolute_file_path)
        flash("File not found", category="danger")
        return "", 404

    df = pd.read_excel(bulk_ship.absolute_file_path)
    excel_html = df.to_html(
        index=False,
    )

    return render_template("bulk_ship/modal_view.html", excel_html=excel_html)


@bulk_ship_bp.route("/<uuid>/delete", methods=["DELETE"])
@login_required
@role_required(
    [s.UserRole.ADMIN.value, s.UserRole.WAREHOUSE_MANAGER.value], has_bulk_ship=True
)
def delete(uuid: str):
    bulk_ship = db.session.scalar(sa.select(m.BulkShip).where(m.BulkShip.uuid == uuid))
    if not bulk_ship or bulk_ship.is_deleted:
        log(log.ERROR, "Bulk ship not found [%s]", uuid)
        flash("Bulk ship not found", category="danger")
        return "", 404

    bulk_ship.is_deleted = True
    bulk_ship.name = f"{bulk_ship.name} - Deleted at {datetime.now()}"
    db.session.commit()

    flash("Bulk ship deleted successfully", category="success")
    return "", 204
