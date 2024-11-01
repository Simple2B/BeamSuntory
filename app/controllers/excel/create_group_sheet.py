from flask_login import current_user
from openpyxl import Workbook
from app import models as m, db
import sqlalchemy as sa
from openpyxl.utils import get_column_letter
from app import schema as s
from openpyxl.worksheet.datavalidation import DataValidation

from app.views.utils import create_match


def create_group_sheet(wb: Workbook, main_sheet, group_names):
    groups_sheet = wb.create_sheet(title="Groups")

    products = db.session.scalars(
        sa.select(m.Product).where(
            m.Product.warehouse_products.any(
                sa.and_(
                    m.WarehouseProduct.available_quantity != 0,
                    m.WarehouseProduct.group.has(m.Group.name.in_(group_names)),
                )
            )
        )
    ).all()
    product_count = len(products)
    # product_count = db.session.scalar(
    #     sa.select(sa.func.count())
    #     .select_from(m.Product)
    #     .where(
    #         m.Product.warehouse_products.any(
    #             sa.and_(
    #                 m.WarehouseProduct.available_quantity != 0,
    #                 m.WarehouseProduct.group.has(m.Group.name.in_(group_names)),
    #             )
    #         )
    #     )
    # )

    for col, product in enumerate(products, start=1):
        col_letter = get_column_letter(col)
        groups_sheet[f"{col_letter}1"] = product.SKU
        wh_products = [
            wh_prod
            for wh_prod in product.warehouse_products
            if wh_prod.available_quantity != 0
            and wh_prod.group.name != s.Events.name.value
            and wh_prod.group.name in current_user.user_group_names
        ]
        for row, wh_product in enumerate(wh_products, start=2):
            groups_sheet[f"{col_letter}{row}"] = (
                f"{wh_product.group.name} ({wh_product.available_quantity})"
            )

    # Add dropdowns for product SKUs
    last_let = get_column_letter(product_count)
    country_dv = DataValidation(
        type="list",
        formula1=f"{groups_sheet.title}!$A$1:${last_let}$1",
        allow_blank=True,
        showDropDown=False,
    )
    main_sheet.add_data_validation(country_dv)
    country_dv.add("A2:A100")

    # Create the named ranges
    for row in range(2, 100):
        groups_dv = DataValidation(
            type="list",
            formula1=f"OFFSET({groups_sheet.title}!A1,1,{create_match('A' + str(row),groups_sheet.title, product_count)},50,1)",
            showDropDown=False,
            allow_blank=False,
            showErrorMessage=True,
        )
        groups_dv.error = "Please select a valid group"
        groups_dv.errorTitle = "Invalid group"
        main_sheet.add_data_validation(groups_dv)
        groups_dv.add(f"B{row}")
