from flask_login import current_user
from app import models as m, db
from openpyxl import Workbook
import sqlalchemy as sa
from openpyxl.utils import get_column_letter
from app import schema as s
from openpyxl.worksheet.datavalidation import DataValidation

from app.views.utils import create_match


def create_master_group_sheet(wb: Workbook, main_sheet, group_names):
    master_group_sheet = wb.create_sheet(title="Master_Groups")

    master_groups = db.session.scalars(sa.select(m.MasterGroup)).all()

    # groups_count = len(
    #     [
    #         group
    #         for group in master_group.groups
    #         if group.name not in s.Events.name.value
    #         and group.name in current_user.user_group_names
    #     ]
    # )

    for col, master_group in enumerate(master_groups, start=1):
        mg_col_letter = get_column_letter(col)
        master_group_sheet[f"{mg_col_letter}1"] = master_group.name
        groups = [
            group
            for group in master_group.groups
            if group.name not in s.Events.name.value
            and group.name in current_user.user_group_names
        ]
        for row, group in enumerate(groups, start=2):
            master_group_sheet[f"{mg_col_letter}{row}"] = group.name

    # Add dropdowns for master groups
    last_let = get_column_letter(len(master_groups))
    master_group_dv = DataValidation(
        type="list",
        formula1=f"{master_group_sheet.title}!$A$1:${last_let}$1",
        allow_blank=True,
        showDropDown=False,
    )
    main_sheet.add_data_validation(master_group_dv)
    master_group_dv.add("C2:C100")

    # create a named range for the master groups
    for row in range(2, 100):
        groups_dv = DataValidation(
            type="list",
            formula1=f"OFFSET({master_group_sheet.title}!A1,1,{create_match('C' + str(row),master_group_sheet.title, len(master_groups))},100,1)",
            showDropDown=False,
            allow_blank=False,
            showErrorMessage=True,
        )
        groups_dv.error = "Please select a valid group"
        groups_dv.errorTitle = "Invalid group"
        main_sheet.add_data_validation(groups_dv)
        groups_dv.add(f"D{row}")

        quantity_dv = DataValidation(
            type="custom",
            formula1=f'AND(ISNUMBER(E{row}), E{row} <= MID(B{row}, FIND("(", B{row}) + 1, FIND(")", B{row}) - FIND("(", B{row}) - 1) * 1)',
            showDropDown=False,
            allow_blank=False,
            showErrorMessage=True,
            error="Please enter a number not greater than the value in column B.",
            errorTitle="Invalid input",
        )
        main_sheet.add_data_validation(quantity_dv)
        quantity_dv.add(f"E{row}")
