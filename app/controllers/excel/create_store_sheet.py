from app import models as m, db
import sqlalchemy as sa
from openpyxl.utils import get_column_letter


def create_store_sheet(wb, main_sheet, GROUP_NAMES):
    store_categories = db.session.scalars(
        sa.select(m.StoreCategory).where(m.StoreCategory.name.not_ilike("%Event%"))
    )
    store_categories_sheet = wb.create_sheet(title="Stores")
    for col, store_category in enumerate(store_categories, start=1):
        col_letter = get_column_letter(col)
        store_categories_sheet[f"{col_letter}1"] = store_category.name
        for row, store in enumerate(store_category.stores, start=2):
            store_categories_sheet[f"{col_letter}{row}"] = store.store_name
